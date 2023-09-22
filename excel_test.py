from openpyxl import load_workbook

excel_file = "test.xlsx"
wb = load_workbook(excel_file, data_only=True)
sh = wb['Sheet1']

for i in range(1, sh.max_row + 1):
    for j in range(1, sh.max_column + 1):
        print(f'row {i}', f'col {j}', f'color {sh.cell(i, j).fill.fgColor.rgb}', end= ' ')
